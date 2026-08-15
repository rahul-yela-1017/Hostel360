import csv
import io
import zipfile
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload
from ..database import get_db
from ..deps import require_roles
from ..models import Bed, BedStatus, Role, Room, RoomStatus, Student, User
from ..services import audit, room_scope_for

router = APIRouter(prefix="/reports", tags=["Reports"])


def selected_rooms(db: Session, hostel_id: int, room_from: int, room_to: int, status: str):
    if room_from > room_to:
        raise HTTPException(400, "Room From must not exceed Room To")
    query = select(Room).where(Room.hostel_id == hostel_id, Room.room_number.between(room_from, room_to))
    status_map = {"occupied": RoomStatus.FULL, "partially occupied": RoomStatus.PARTIALLY_OCCUPIED, "empty": RoomStatus.EMPTY}
    if status.lower() != "all":
        mapped = status_map.get(status.lower())
        if not mapped: raise HTTPException(400, "Invalid room status filter")
        query = query.where(Room.status == mapped)
    return db.scalars(query.options(selectinload(Room.beds).selectinload(Bed.student).joinedload(Student.user))
                      .order_by(Room.room_number)).all()


def rows_for(room: Room):
    for bed in room.beds:
        student = bed.student
        yield [room.room_number, bed.bed_number, student.user.name if student else "",
               student.roll_no if student else "", "Occupied" if student else "Vacant"]


def csv_bytes(rooms: list[Room], include_room=True) -> bytes:
    stream = io.StringIO(newline="")
    writer = csv.writer(stream)
    writer.writerow((["Room Number"] if include_room else []) + ["Bed Number", "Student Name", "Roll Number", "Status"])
    for room in rooms:
        for row in rows_for(room):
            writer.writerow(row if include_room else row[1:])
    return stream.getvalue().encode("utf-8-sig")


def xlsx_bytes(rooms: list[Room]) -> bytes:
    wb = Workbook()
    ws = wb.active; ws.title = "Complete Hostel Room List"
    headers = ["Room Number", "Bed Number", "Student Name", "Roll Number", "Status"]
    ws.append(headers)
    for room in rooms:
        for row in rows_for(room): ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="172B4D")
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    for key, width in {"A": 15, "B": 13, "C": 28, "D": 20, "E": 15}.items(): ws.column_dimensions[key].width = width
    output = io.BytesIO(); wb.save(output); return output.getvalue()


@router.get("/rooms")
def room_report_preview(
    room_from: int = Query(0, ge=0, le=430), room_to: int = Query(430, ge=0, le=430),
    status: str = "All", db: Session = Depends(get_db), actor: User = Depends(require_roles(Role.ADMIN, Role.WARDEN)),
):
    hostel_id = room_scope_for(actor) or 1
    rooms = selected_rooms(db, hostel_id, room_from, room_to, status)
    return {"room_count": len(rooms), "bed_count": len(rooms)*4,
            "occupied": sum(1 for r in rooms for b in r.beds if b.status == BedStatus.OCCUPIED),
            "rooms": [{"room_number": r.room_number, "status": r.status.value,
                       "occupied": sum(1 for b in r.beds if b.status == BedStatus.OCCUPIED)} for r in rooms[:100]]}


@router.get("/rooms/download")
def download_room_report(
    room_from: int = Query(0, ge=0, le=430), room_to: int = Query(430, ge=0, le=430),
    status: str = "All", format: str = Query("zip", pattern="^(csv|xlsx|zip)$"),
    db: Session = Depends(get_db), actor: User = Depends(require_roles(Role.ADMIN, Role.WARDEN)),
):
    hostel_id = room_scope_for(actor) or 1
    rooms = selected_rooms(db, hostel_id, room_from, room_to, status)
    audit(db, actor.id, "ROOM_REPORT_DOWNLOADED", "report", None,
          {"from": room_from, "to": room_to, "status": status, "format": format, "rooms": len(rooms)})
    db.commit()

    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    if format == "csv":
        data, media, filename = csv_bytes(rooms), "text/csv", f"Complete_Hostel_Room_List_{stamp}.csv"
    elif format == "xlsx":
        data, media, filename = xlsx_bytes(rooms), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"Hostel_Room_Report_{stamp}.xlsx"
    else:
        archive = io.BytesIO()
        with zipfile.ZipFile(archive, "w", zipfile.ZIP_DEFLATED) as zf:
            for room in rooms:
                zf.writestr(f"Room_{room.room_number:03d}.csv", csv_bytes([room], include_room=False))
            zf.writestr("Complete_Hostel_Room_List.csv", csv_bytes(rooms))
            zf.writestr("Complete_Hostel_Room_List.xlsx", xlsx_bytes(rooms))
            zf.writestr("README.txt", f"Smart Hostel room occupancy report\nRooms: {room_from} to {room_to}\nFilter: {status}\nGenerated: {datetime.now().isoformat()}\n")
        data, media, filename = archive.getvalue(), "application/zip", f"Hostel_Room_Report_{stamp}.zip"
    return StreamingResponse(io.BytesIO(data), media_type=media,
                             headers={"Content-Disposition": f'attachment; filename="{filename}"',
                                      "Cache-Control": "no-store"})
